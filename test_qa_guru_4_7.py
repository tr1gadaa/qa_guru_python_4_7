import zipfile
import PyPDF2
from PyPDF2 import PdfReader
import pytest
from openpyxl import load_workbook
from openpyxl import workbook
import csv
import os


@pytest.fixture()
def create_zip_file():
    filenames = [os.path.abspath("example_csv.csv"), os.path.abspath("example_pdf.pdf"),
                 os.path.abspath("example_xslsx.xlsx")]

    with zipfile.ZipFile("resources/archive.zip", 'w') as archive:
        for filename in filenames:
            if os.path.exists(filename):
                archive.write(filename, os.path.basename(filename))
            else:
                print(f'Файл {filename} не найден.')
        archive.close()



def test_check_pdf(create_zip_file):
    with zipfile.ZipFile("resources/archive.zip", "r") as archive:
        with archive.open("example_pdf.pdf", "r") as file_pdf:
            reader = PyPDF2.PdfReader(file_pdf)
            number_of_pages = len(reader.pages)
            assert number_of_pages == 2


def test_check_csv(create_zip_file):
    with zipfile.ZipFile("resources/archive.zip", "r") as archive:
        with archive.open("example_csv.csv", "r") as file_csv:
            row = [row for row in file_csv]
            assert row[0] == b'Login email;Identifier;First name;Last name\n'
            assert row[1] == b'laura@example.com;2070;Laura;Grey\n'


def test_check_xlsx(create_zip_file):
    workbook = load_workbook("example_xslsx.xlsx")
    sheet = workbook.active
    text_value = sheet.cell(row=4, column=5).value
    with zipfile.ZipFile("resources/archive.zip", "r") as archive:
        with archive.open("example_xslsx.xlsx", "r") as file_xlsx:
            workbook = load_workbook(file_xlsx)
            reader = workbook
            file_xlsx_sheet = reader.active
            file_xlsx_value = file_xlsx_sheet.cell(row=4, column=5).value
            assert text_value == file_xlsx_value
